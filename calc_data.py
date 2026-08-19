#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
李家村门店业绩看板 —— 明细复算引擎
直接从用友云导出的销售明细 TSV 复现《李家村销售》表的全部 SUMIFS 口径，生成 data.json。

为什么要复算？
    表格里的指标全是 SUMIFS 公式。脚本写完 XS/RXS 明细后，公式的「缓存值」
    要等 WPS/Excel 打开才会刷新。看板不能等人开表格，所以这里按同一套口径
    自己算一遍，做到「抓完数 → 看板立刻是新的」。

口径来源：逐格导出《李家村销售》表的真实公式，1:1 复现，不做任何自创逻辑。
    · 任务量  → 读「8月任务」表常量（纯手工填写，不受公式缓存影响）
    · 完成量  → 由明细 TSV 复算
    · 手工项  → 乐回收(常量) / 太力回收(独立表) / 绩效(个人表) 直接读，它们不依赖明细

用法：
    python calc_data.py <明细.tsv> [--xlsx 路径] [--day YYYY-MM-DD] [-o data.json]
"""
import sys, os, re, json, csv, datetime, argparse, calendar
import openpyxl

# ---------- 明细列（与用友云导出、XS/RXS 完全一致的 19 列） ----------
# Excel 列字母 -> 0-based 下标
C = {c: i for i, c in enumerate("ABCDEFGHIJKLMNOPQRS")}
HEADERS = ["出库单号", "单据类型", "出库日期", "商品分类", "商品sku分类", "商品SKU编码",
           "商品名称", "入库属性", "数量", "单价", "原价", "折扣价", "金额", "毛利",
           "SO激励", "业务员", "库区", "销售出库单门店", "销售成本"]

PEOPLE_ORDER = ["邵乐乐", "杨丽华", "李泽", "陈超磊", "张博晨"]
TASK_PEOPLE = ["邵乐乐", "杨丽华", "李泽", "陈超磊"]          # 8月任务表里有任务的 4 人
TASK_ROW = {"邵乐乐": 4, "杨丽华": 5, "李泽": 6, "陈超磊": 7}   # 8月任务表行号

# 8月任务表列：C=销售额 D=毛利 E=手机 F=增值 G=积分 H=PC I=平板 J=音频 K=穿戴 L=HD M=摄影 N=考核机
TASK_COL = {"销额": 3, "毛利": 4, "手机": 5, "增值": 6, "积分": 7,
            "PC": 8, "平板": 9, "音频": 10, "穿戴": 11, "HD": 12,
            "摄影课": 13, "考核机型": 14}

# 考核机型 SKU 前缀（李家村销售!S14 数组公式原样搬运）
KHJX_SKU = ["01.001.010.00*", "01.001.011.002*", "01.001.012.00*", "01.001.013.0*",
            "01.001.031.002*", "01.001.032.002*", "01.001.043.002*", "01.001.044.002*",
            "01.001.001.0*", "01.001.002.0*", "01.001.003*"]
# care+ 里计入的星联优享档位（G14 数组公式原样搬运）
XLYX = ["星联优享-499", "星联优享-699", "星联优享-999"]


# ============================ 基础工具 ============================
def num(v):
    """明细里的数字：带千分位、可能为空。"""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("¥", "").replace("￥", "")
    if not s or s.startswith("#"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def wild(text, pat):
    """Excel 通配符匹配：* 任意长度，? 单字符，整体锚定，不区分大小写。"""
    rx = "".join(".*" if ch == "*" else "." if ch == "?" else re.escape(ch) for ch in pat)
    return re.fullmatch(rx, text, re.IGNORECASE) is not None


def hit(cell, cond):
    """复现 SUMIFS 单个条件。cond 支持 '>0' 这类比较式，和带通配符的文本。"""
    if isinstance(cond, (int, float)):
        return num(cell) == float(cond)
    cond = str(cond)
    m = re.fullmatch(r"(>=|<=|<>|>|<|=)\s*(-?[\d.]+)", cond)
    if m:
        op, rhs = m.group(1), float(m.group(2))
        x = num(cell)
        return {">": x > rhs, "<": x < rhs, ">=": x >= rhs,
                "<=": x <= rhs, "=": x == rhs, "<>": x != rhs}[op]
    s = "" if cell is None else str(cell).strip()
    if any(ch in cond for ch in "*?"):
        return wild(s, cond)
    return s.lower() == cond.strip().lower()


def sumifs(rows, sum_col, *conds):
    """SUMIFS(明细[sum_col], 条件对...)。conds 形如 ('P','邵乐乐'), ('D','01手机')"""
    total = 0.0
    for r in rows:
        if all(hit(r[C[col]], cond) for col, cond in conds):
            total += num(r[C[sum_col]])
    return total


def sumifs_any(rows, sum_col, base_conds, col, patterns):
    """数组常量版：SUM(SUMIFS(..., {p1,p2,...}))。同一行命中多个模式会重复计入（与 Excel 一致）。"""
    total = 0.0
    for p in patterns:
        total += sumifs(rows, sum_col, *base_conds, (col, p))
    return total


def roundup(x, digits=0):
    """Excel ROUNDUP：远离 0 取整。"""
    import math
    f = 10 ** digits
    return math.ceil(x * f) / f if x >= 0 else math.floor(x * f) / f


def div(a, b):
    """除法，分母为 0 返回 None（对应表格里的 #DIV/0!）。"""
    try:
        return a / b if b else None
    except (TypeError, ZeroDivisionError):
        return None


# ============================ 读明细 ============================
def load_tsv(path):
    # 兼容用友导出的 GBK/UTF-8 混合文件：先按 utf-8-sig 读，失败再回退 gbk
    raw = open(path, "rb").read()
    try:
        txt = raw.decode("utf-8-sig")
    except Exception:
        txt = raw.decode("gbk", "replace")
    rd = list(csv.reader(txt.splitlines(), delimiter="\t"))
    if not rd:
        sys.exit("❌ 明细文件是空的")
    head = [h.strip() for h in rd[0]]
    if head[:len(HEADERS)] != HEADERS:
        sys.exit(f"❌ 明细表头与预期不符\n  期望: {HEADERS}\n  实际: {head}")
    rows = [r + [""] * (len(HEADERS) - len(r)) for r in rd[1:] if any(x.strip() for x in r)]
    return rows


def day_of(r):
    return str(r[C["C"]]).strip()[:10]


# ============================ 读表格里的手工项 ============================
def load_manual(xlsx):
    """任务量、乐回收、太力回收、绩效——这几样不来自明细，从表格取。"""
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    wbf = openpyxl.load_workbook(xlsx, data_only=False)

    # --- 任务量：8月任务表，纯常量 ---
    tk = wbf[[s for s in wbf.sheetnames if s.endswith("月任务")][0]]
    tasks = {}
    for name, row in TASK_ROW.items():
        tasks[name] = {k: num(tk.cell(row, col).value) for k, col in TASK_COL.items()}
    tasks["张博晨"] = {k: 0.0 for k in TASK_COL}          # 张博晨不背任务

    # --- 乐回收：李家村销售 T/V 列常量 ---
    ws = wbf["李家村销售"]
    P2_ROWS = {"邵乐乐": 14, "杨丽华": 15, "李泽": 16, "陈超磊": 17, "张博晨": 18}
    lehui = {}
    for name, row in P2_ROWS.items():
        lehui[name] = {"orders": num(ws.cell(row, 20).value),
                       "amount": num(ws.cell(row, 21).value),
                       "增值":   num(ws.cell(row, 22).value)}

    # --- 太力回收：独立表，按业务员汇总（不依赖明细，实时算） ---
    tl = wb["太力回收"]
    taili = {n: {"orders": 0.0, "amount": 0.0} for n in PEOPLE_ORDER}
    for r in range(2, tl.max_row + 1):
        state = str(tl.cell(r, 10).value or "").strip()       # J 回收单状态
        who = str(tl.cell(r, 22).value or "").strip()         # V 销售员姓名
        if who in taili:
            if state == "已付款":
                taili[who]["orders"] += num(tl.cell(r, 27).value)   # AA 数量
            taili[who]["amount"] += num(tl.cell(r, 28).value)       # AB 回收价（SUMIF 不筛状态）
    for n in taili:
        taili[n]["增值"] = taili[n]["amount"] * 0.14

    # --- 绩效：各人表 L18（=SUM(L4:L17)，与明细无关） ---
    # 注意：原表《李家村销售》AP 列引用错位（邵乐乐行→杨丽华表、杨丽华行→李泽表…），
    #       且邵乐乐表 L 列公式为 #REF!。这里按「每人读自己的表」的正确逻辑取值，
    #       取不到就给 None（看板显示「—」），不拿别人的数顶替。
    perf = {}
    for n in PEOPLE_ORDER:
        if n not in wb.sheetnames:
            perf[n] = None
            continue
        v = wb[n].cell(18, 12).value
        if v is None or (isinstance(v, str) and (v.strip() == "" or v.startswith("#"))):
            perf[n] = None                       # #REF! / 空 —— 数据坏了，如实标记
        else:
            perf[n] = num(v)

    # --- 表头标签（静态文本） ---
    lab_day = [ws.cell(26, c).value for c in range(2, 16)]
    lab_gap = [ws.cell(36, c).value for c in range(2, 16)]
    return tasks, lehui, taili, perf, lab_day, lab_gap


# ============================ 复算：业绩考核 ============================
def calc_perf(xs, name, task):
    P = ("P", name)
    毛利 = sumifs(xs, "N", P)
    手机 = sumifs(xs, "I", P, ("D", "01手机"))
    PC = sumifs(xs, "I", P, ("D", "05电脑"))
    平板 = sumifs(xs, "I", P, ("D", "06平板电脑"))
    穿戴 = sumifs(xs, "I", P, ("D", "08穿戴"))
    音频 = sumifs(xs, "I", P, ("D", "07音频"))
    HD = sumifs(xs, "I", P, ("F", "12*"))
    销额 = sumifs(xs, "M", P)

    def blk(t, d, on_zero=None):
        # on_zero：任务为 0 时的完成率。HD 列公式是 IFERROR(...,"100%")，照搬。
        return {"task": t, "done": d, "gap": d - t,
                "rate": div(d, t) if t else on_zero}

    return {
        "毛利":     blk(task["毛利"], 毛利),
        "手机":     blk(task["手机"], 手机),
        "PC":       blk(task["PC"], PC),
        "平板":     blk(task["平板"], 平板),
        "穿戴":     blk(task["穿戴"], 穿戴),
        "音频":     blk(task["音频"], 音频),
        "HD":       blk(task["HD"], HD, on_zero=1.0),
        "智慧办公": blk(task["PC"] + task["平板"], PC + 平板),
        "音频穿戴": blk(task["穿戴"] + task["音频"], 穿戴 + 音频),
        "销额":     blk(task["销额"], 销额),
    }


# ============================ 复算：品类销售明细（点击展开用）============================
# 复用 calc_perf 的归类口径，把每一行出库明细归入对应品类，供前端「品类达成」点击下钻。
DET_CATS = {
    "手机":     lambda r: r[C["D"]] == "01手机",
    "PC":       lambda r: r[C["D"]] == "05电脑",
    "平板":     lambda r: r[C["D"]] == "06平板电脑",
    "穿戴":     lambda r: r[C["D"]] == "08穿戴",
    "音频":     lambda r: r[C["D"]] == "07音频",
    "HD":       lambda r: (r[C["F"]] or "").startswith("12"),
    "智慧办公": lambda r: r[C["D"]] in ("05电脑", "06平板电脑"),
    "音频穿戴": lambda r: r[C["D"]] in ("08穿戴", "07音频"),
    "增值":     lambda r: "增值" in (r[C["D"]] or ""),
}

def row_detail(r):
    amt = num(r[C["M"]])   # 金额
    pf  = num(r[C["N"]])   # 毛利
    gpr = round(pf / amt, 4) if amt else None   # 毛利率 = 毛利 / 金额
    return {
        "date":      r[C["C"]],
        "name":      r[C["G"]],
        "qty":       num(r[C["I"]]),
        "origPrice": num(r[C["K"]]),   # 原价
        "discPrice": num(r[C["L"]]),   # 折扣价
        "amount":    amt,
        "profit":    pf,
        "gpr":       gpr,              # 毛利率（自算）
        "so":        (r[C["O"]] or "").strip(),  # SO激励
        "cost":      num(r[C["S"]]),   # 销售成本
        "emp":       r[C["P"]],
        "sku":       r[C["F"]],
    }

def build_details(xs):
    out = {}
    for cat, cond in DET_CATS.items():
        rows = [row_detail(r) for r in xs if cond(r)]
        rows.sort(key=lambda x: (x["date"] or ""), reverse=True)
        out[cat] = rows
    return out


# ============================ 复算：全科生 / 增值 ============================
def calc_qcs(xs, name, task, perf, lehui, taili):
    P = ("P", name)
    手机 = perf["手机"]["done"]
    智慧办公 = perf["智慧办公"]["done"]
    穿戴 = perf["穿戴"]["done"]
    毛利 = perf["毛利"]["done"]
    销额 = perf["销额"]["done"]

    # 电信积分
    积分完成 = sumifs(xs, "M", P, ("D", "10运营商业务"))
    积分任务 = task["积分"]

    # 会员搭售（care+）
    care = sumifs(xs, "I", P, ("G", "*Care*"), ("N", ">0")) + \
           sumifs_any(xs, "I", [P], "G", XLYX)
    care_gap = care - roundup(手机 * 0.30)

    # 回收搭售
    回收 = sumifs(xs, "I", P, ("G", "*回收*")) + taili[name]["orders"] + lehui[name]["orders"]
    回收_gap = 回收 - roundup(手机 * 0.20)

    # 贴膜
    贴膜 = sumifs(xs, "I", P, ("G", "*膜*"), ("N", ">0")) + sumifs(xs, "I", P, ("G", "*贴膜套包"))
    贴膜基数 = 手机 + 智慧办公 + 穿戴
    贴膜_gap = 贴膜 - roundup(贴膜基数 * 0.50)

    # 考核机型：完成 = 各 SKU 前缀命中之和；表格里 S 列存的是「缺口 = 完成 - 任务」
    考核完成 = sumifs_any(xs, "I", [P], "F", KHJX_SKU)
    考核任务 = task["考核机型"]

    # 增值：明细增值毛利 + 太力增值 + 乐回收金额(U列) + 电信积分×4
    # ⚠️ 口径来自《李家村销售》真实公式：增值 = SUMIFS(明细增值) + Y(太力增值) + U(乐回收金额) + C(电信积分)×4
    #    之前错把「乐回收金额(U)」写成「乐回收增值(V，常空)」、积分倍数写成 ×3，导致看板增值偏低。2026-08-13 修正。
    增值完成 = sumifs(xs, "N", P, ("D", "*增值*")) + taili[name]["增值"] + lehui[name]["amount"] + 积分完成 * 4
    增值任务 = task["增值"]

    # 健康度
    优惠券 = sumifs(xs, "L", P)

    # 星联会员
    优享 = sumifs(xs, "I", P, ("G", "*会员*")) + sumifs(xs, "I", P, ("G", "星联优享*"))
    尊享 = sumifs(xs, "I", P, ("G", "*储值*")) + sumifs(xs, "I", P, ("G", "星联尊享*"))

    return {
        "电信积分":   {"task": 积分任务, "done": 积分完成,
                     "gap": 积分完成 - 积分任务, "rate": div(积分完成, 积分任务)},
        "会员搭售率": {"terminal": 手机, "care": care, "gap": care_gap, "rate": div(care, 手机)},
        "回收搭售率": {"orders": 回收, "gap": 回收_gap, "rate": div(回收, 手机)},
        "贴膜率":     {"orders": 贴膜, "gap": 贴膜_gap, "rate": div(贴膜, 贴膜基数)},
        "考核机型":   {"task": 考核任务, "done": 考核完成, "gap": 考核完成 - 考核任务},
        "乐回收":     dict(lehui[name]),
        "太力回收":   {"orders": taili[name]["orders"], "amount": taili[name]["amount"],
                     "增值": taili[name]["增值"]},
        "增值":       {"task": 增值任务, "done": 增值完成,
                     "gap": 增值完成 - 增值任务, "rate": div(增值完成, 增值任务)},
        "健康度":     {"coupon": 优惠券, "ratio": div(优惠券, 毛利),
                     "grossMargin": div(毛利, 销额), "增值率": div(增值完成, 销额)},
        "星联会员":   {"优享": 优享, "尊享": 尊享, "合计": 优享 + 尊享},
    }


# ============================ 复算：当日达成 ============================
def calc_daily(rxs, name, labels):
    P = ("P", name)
    # 严格对齐《李家村销售》sheet「今日达成」区块(行28-33)的 SUMIFS 公式：
    #   增值=毛利列且(分类∈{*增值*,*运营商*})；会员=Care+会员+星联优享*；
    #   贴膜=*膜*&毛利>0+贴膜套包+会员；摄影课=*大师课*&毛利>0；滞销=KHJX_SKU数组
    v = {
        "手机":     sumifs(rxs, "I", P, ("F", "01.001*")),
        "毛利":     sumifs(rxs, "N", P),
        "增值":     sumifs(rxs, "N", P, ("D", "*增值*")) + sumifs(rxs, "N", P, ("D", "*运营商*")),
        "智慧办公": sumifs(rxs, "I", P, ("F", "05.001*")) + sumifs(rxs, "I", P, ("F", "06.001*")),
        "音频穿戴": sumifs(rxs, "I", P, ("F", "08.001*")) + sumifs(rxs, "I", P, ("F", "07.001*")),
        "HD":       sumifs(rxs, "I", P, ("F", "12*")),
        "会员":     sumifs(rxs, "I", P, ("G", "*Care*")) + sumifs(rxs, "I", P, ("G", "*会员*")) + sumifs(rxs, "I", P, ("G", "星联优享*")),
        "回收":     sumifs(rxs, "I", P, ("G", "*回收*")),
        "贴膜":     sumifs(rxs, "I", P, ("G", "*膜*"), ("N", ">0")) + sumifs(rxs, "I", P, ("G", "*贴膜套包")) + sumifs(rxs, "I", P, ("G", "*会员*")),
        "电信积分": sumifs(rxs, "M", P, ("E", "*入网*")),
        "滞销":     sumifs_any(rxs, "I", [P], "F", KHJX_SKU),
        "摄影课":   sumifs(rxs, "I", P, ("G", "*大师课*"), ("N", ">0")),
        "优享/会员": sumifs(rxs, "I", P, ("G", "*新自由*")) + sumifs(rxs, "I", P, ("G", "星联优享*")),
    }
    # 严格按表格「今日达成」品类标签(B26:N26)输出（含原被误排除的「摄影课」）
    out = {k: v.get(k, 0.0) for k in labels if k}
    out["销额"] = sumifs(rxs, "M", P)   # 表格「今日达成」无销额行，但日报总览/板块依赖，保留
    return out


# ============================ 复算：每日缺口 ============================
def remain_days(base):
    """剩余天数 = (月末 - 当日) - MAX(0, 4 - INT((DAY-1)/7))，与表格公式一致。"""
    last = calendar.monthrange(base.year, base.month)[1]
    return max(1, (last - base.day) - max(0, 4 - (base.day - 1) // 7))


def calc_gap(perf, qcs, rd, labels):
    """每日缺口 = ROUNDUP(该项缺口 / 剩余天数)。缺口为负，结果即「每天还差多少」。"""
    src = {
        "手机":      perf["手机"]["gap"],
        "毛利":      perf["毛利"]["gap"],
        "增值":      qcs["增值"]["gap"],
        "智慧办公":  perf["智慧办公"]["gap"],
        "音频穿戴":  perf["音频穿戴"]["gap"],
        "HD":        perf["HD"]["gap"],
        "Care+":     qcs["会员搭售率"]["gap"],
        "回收":      qcs["回收搭售率"]["gap"],
        "贴膜":      qcs["贴膜率"]["gap"],
        "电信积分":  qcs["电信积分"]["gap"],
        "滞销":      qcs["考核机型"]["gap"],
    }
    out = {}
    for k in labels:
        if k and k != "摄影课" and k in src:
            out[k] = roundup(src[k] / rd)
    return out


# ============================ 汇总 ============================
def total_perf(people):
    ref = people[PEOPLE_ORDER[0]]["performance"]
    keys = [k for k, v in ref.items() if isinstance(v, dict)]   # 绩效是标量，单独汇总
    out = {}
    for k in keys:
        t = sum(people[n]["performance"][k]["task"] for n in PEOPLE_ORDER)
        d = sum(people[n]["performance"][k]["done"] for n in PEOPLE_ORDER)
        out[k] = {"task": t, "done": d, "gap": d - t, "rate": div(d, t)}
    return out


def total_qcs(people, sp):
    S = lambda k, f: sum(people[n]["qcs"][k][f] for n in PEOPLE_ORDER)
    手机, 智慧办公, 穿戴 = sp["手机"]["done"], sp["智慧办公"]["done"], sp["穿戴"]["done"]
    毛利, 销额 = sp["毛利"]["done"], sp["销额"]["done"]
    care, 回收, 贴膜 = S("会员搭售率", "care"), S("回收搭售率", "orders"), S("贴膜率", "orders")
    贴膜基数 = 手机 + 智慧办公 + 穿戴
    jf_t, jf_d = S("电信积分", "task"), S("电信积分", "done")
    kh_t, kh_d = S("考核机型", "task"), S("考核机型", "done")
    zz_t, zz_d = S("增值", "task"), S("增值", "done")
    优惠券 = S("健康度", "coupon")
    # 合计行的三个「缺口」不是各人相加，而是拿合计量重算（表格 H19/K19/N19 原样搬运）：
    #   care  = G19 - ROUNDUP(F19*30%)   回收 = J19 - ROUNDUP(G9*20%)
    #   贴膜  = M19 - ROUNDUP(G9*50%)    ← 合计行基数只取手机，与个人行口径不同
    return {
        "电信积分":   {"task": jf_t, "done": jf_d, "gap": jf_d - jf_t, "rate": div(jf_d, jf_t)},
        "会员搭售率": {"terminal": 手机, "care": care,
                     "gap": care - roundup(手机 * 0.30), "rate": div(care, 手机)},
        "回收搭售率": {"orders": 回收, "gap": 回收 - roundup(手机 * 0.20), "rate": div(回收, 手机)},
        "贴膜率":     {"orders": 贴膜, "gap": 贴膜 - roundup(手机 * 0.50), "rate": div(贴膜, 贴膜基数)},
        "考核机型":   {"task": kh_t, "done": kh_d, "gap": kh_d - kh_t},
        "乐回收":     {f: S("乐回收", f) for f in ("orders", "amount", "增值")},
        "太力回收":   {f: S("太力回收", f) for f in ("orders", "amount", "增值")},
        "增值":       {"task": zz_t, "done": zz_d, "gap": zz_d - zz_t, "rate": div(zz_d, zz_t)},
        "健康度":     {"coupon": 优惠券, "ratio": div(优惠券, 毛利),
                     "grossMargin": div(毛利, 销额), "增值率": div(zz_d, 销额)},
        "星联会员":   {f: S("星联会员", f) for f in ("优享", "尊享", "合计")},
    }


# ============================ 店长洞察 ============================
# 看板品类 → 明细筛选条件（用于算最后成交日 / 断销天数）
CAT_FILTER = {
    "手机":  ("D", "01手机"),
    "PC":    ("D", "05电脑"),
    "平板":  ("D", "06平板电脑"),
    "穿戴":  ("D", "08穿戴"),
    "音频":  ("D", "07音频"),
    "HD":    ("F", "12*"),
}
CAT_UNIT = {"手机": "台", "PC": "台", "平板": "台", "穿戴": "件", "音频": "件", "HD": "台"}
CAT_ALIAS = {"PC": "电脑", "HD": "HD 智慧屏"}


def last_sold(xs, cond, name=None):
    """某品类（可限定业务员）最后一次成交的日期。没卖过返回 None。"""
    col, pat = cond
    best = None
    for r in xs:
        if not hit(r[C[col]], pat):
            continue
        if name and r[C["P"]].strip() != name:
            continue
        if num(r[C["I"]]) <= 0:          # 只认正向出货，退货不算动销
            continue
        d = day_of(r)
        if best is None or d > best:
            best = d
    return best


def build_insights(xs, rxs, people, store, ref, tp, rd):
    """把明细 + 复算结果，翻译成店长看得懂的判断和建议。"""

    # ---------- 1. 品类体检 ----------
    cats = []
    for key, cond in CAT_FILTER.items():
        b = store["performance"].get(key) or {}
        task, done = b.get("task", 0), b.get("done", 0)
        rate = b.get("rate") or 0
        ld = last_sold(xs, cond)
        cold = (ref - datetime.date.fromisoformat(ld)).days if ld else None
        # 谁卖过这个品类
        sellers = {}
        for n in TASK_PEOPLE:
            v = people[n]["performance"].get(key, {}).get("done", 0)
            if v:
                sellers[n] = v
        if task <= 0:
            level = "none"
        elif done <= 0:
            level = "danger"                 # 整月零成交
        elif rate < tp * 0.5:
            level = "danger"                 # 进度不到时间的一半
        elif rate < tp * 0.85:
            level = "warn"
        elif rate >= tp:
            level = "good"
        else:
            level = "mid"
        cats.append({
            "key": key,
            "name": CAT_ALIAS.get(key, key),
            "unit": CAT_UNIT.get(key, ""),
            "task": task, "done": done, "rate": rate,
            "lag": rate - tp,
            "level": level,
            "lastDate": ld,
            "coldDays": cold,
            "needPerDay": roundup(max(0, task - done) / rd) if task else 0,
            "sellers": sellers,
            "zeroPeople": [n for n in TASK_PEOPLE
                           if people[n]["performance"].get(key, {}).get("task", 0) > 0
                           and not people[n]["performance"].get(key, {}).get("done", 0)],
        })
    cats.sort(key=lambda c: (c["level"] not in ("danger", "warn"), c["rate"]))

    # ---------- 2. 员工体检 ----------
    W = {"毛利": 0.5, "手机": 0.3}          # 综合分权重：毛利为主，手机为辅
    plist = []
    for n in TASK_PEOPLE:
        pp, pq = people[n]["performance"], people[n]["qcs"]
        r_gross = pp["毛利"].get("rate") or 0
        r_phone = pp["手机"].get("rate") or 0
        r_add = (pq.get("增值") or {}).get("rate") or 0
        score = r_gross * W["毛利"] + r_phone * W["手机"] + r_add * 0.2
        weak = [CAT_ALIAS.get(k, k) for k in CAT_FILTER
                if pp.get(k, {}).get("task", 0) > 0 and (pp[k].get("rate") or 0) < tp * 0.5]
        strong = [CAT_ALIAS.get(k, k) for k in CAT_FILTER
                  if pp.get(k, {}).get("task", 0) > 0 and (pp[k].get("rate") or 0) >= tp]
        plist.append({
            "name": n, "score": score,
            "毛利": {"done": pp["毛利"]["done"], "task": pp["毛利"]["task"], "rate": r_gross},
            "手机": {"done": pp["手机"]["done"], "task": pp["手机"]["task"], "rate": r_phone},
            "增值": {"done": (pq.get("增值") or {}).get("done", 0),
                    "task": (pq.get("增值") or {}).get("task", 0), "rate": r_add},
            "毛利率": (pq.get("健康度") or {}).get("grossMargin"),
            "券占比": (pq.get("健康度") or {}).get("ratio"),
            "strong": strong, "weak": weak,
            "todayGross": people[n]["dailyDone"].get("毛利", 0),
            "todayPhone": people[n]["dailyDone"].get("手机", 0),
        })
    plist.sort(key=lambda x: -x["score"])
    for i, p in enumerate(plist):
        p["rank"] = i + 1
        p["level"] = "good" if p["score"] >= tp else ("warn" if p["score"] >= tp * 0.6 else "danger")

    # ---------- 3. 今日谁卖了谁没卖 ----------
    orders, gross = {}, {}
    for r in rxs:
        n = r[C["P"]].strip()
        orders[n] = orders.get(n, 0) + 1
        gross[n] = gross.get(n, 0.0) + num(r[C["N"]])
    sold = sorted(({"name": n, "orders": orders[n], "gross": gross[n]}
                   for n in orders if n in PEOPLE_ORDER),
                  key=lambda x: -x["gross"])
    idle = [n for n in TASK_PEOPLE if n not in orders]

    # ---------- 4. 自动建议 ----------
    adv = []

    zero_cat = [c for c in cats if c["task"] > 0 and c["done"] <= 0]
    if zero_cat:
        names = "、".join(f"{c['name']}（任务 {c['task']:.0f}{c['unit']}）" for c in zero_cat)
        adv.append({"level": "danger", "icon": "🚨", "title": "整月零成交品类",
                    "body": f"{names} 本月一台没出。建议今天就定人盯：指定专人负责，"
                            f"每天至少推 3 组客户，样机摆到主动线。"})

    cold = [c for c in cats if c["coldDays"] and c["coldDays"] >= 3 and c["done"] > 0]
    if cold:
        s = "、".join(f"{c['name']}（{c['coldDays']}天，最后 {c['lastDate'][5:]}）" for c in cold)
        adv.append({"level": "warn", "icon": "🧊", "title": "断销品类",
                    "body": f"{s} 已经连续多天零动销。先查三件事：样机是否在位、"
                            f"有没有货、店员会不会讲卖点。"})

    behind = [c for c in cats if c["task"] > 0 and 0 < c["rate"] < tp * 0.6]
    if behind:
        s = "；".join(f"{c['name']} 还差 {c['task']-c['done']:.0f}{c['unit']}，"
                      f"日均要 {c['needPerDay']:.0f}{c['unit']}" for c in behind[:3])
        adv.append({"level": "warn", "icon": "📉", "title": f"进度落后（时间已过 {tp:.0%}）",
                    "body": f"{s}。建议把这几项拆到人头，早会点名报进度。"})

    lag_p = [p for p in plist if p["level"] == "danger"]
    if lag_p:
        s = "、".join(f"{p['name']}（毛利 {p['毛利']['rate']:.0%}）" for p in lag_p)
        adv.append({"level": "warn", "icon": "👤", "title": "需要重点帮扶",
                    "body": f"{s} 综合进度明显掉队。别只催结果，先看是客流少、"
                            f"接待量少，还是成交率低——三种病不同药。"})

    if idle:
        adv.append({"level": "warn", "icon": "⏰", "title": "今日还没开单",
                    "body": f"{'、'.join(idle)} 今天暂无上账记录。"
                            f"先确认是没卖还是没及时上账，卖了要立刻补账。"})

    top = plist[0] if plist else None
    if top and top["score"] >= tp:
        adv.append({"level": "good", "icon": "🏆", "title": "本月标杆",
                    "body": f"{top['name']} 综合进度领先（毛利 {top['毛利']['rate']:.0%}"
                            f"、手机 {top['手机']['rate']:.0%}）。"
                            f"让他在早会讲两句怎么谈的，比店长讲管用。"})

    gm = (store["qcs"].get("健康度") or {}).get("grossMargin")
    cp = (store["qcs"].get("健康度") or {}).get("ratio")
    if gm is not None and gm < 0.13:
        adv.append({"level": "warn", "icon": "💰", "title": "毛利率偏低",
                    "body": f"全店毛利率 {gm:.1%}，低于 13% 参考线。"
                            + (f"优惠券占比 {cp:.0%} 偏高，" if cp and cp > 0.3 else "")
                            + "多推增值和配件搭售，比单纯冲机器数划算。"})

    hot = [c for c in cats if c["level"] == "good"]
    if hot:
        adv.append({"level": "good", "icon": "✅", "title": "进度健康",
                    "body": "、".join(f"{c['name']} {c['rate']:.0%}" for c in hot)
                            + " 已跑赢时间进度，保持节奏就行。"})

    return {"categories": cats, "people": plist,
            "today": {"sold": sold, "idle": idle,
                      "totalOrders": len(rxs),
                      "totalGross": sum(gross.values())},
            "advices": adv,
            "timeProgress": tp, "remainDays": rd}


def day_cat(r):
    """当日明细的单行品类（叶子归类，不重叠）：手机/PC/平板/穿戴/音频/HD/增值/其他。"""
    d = str(r[C["D"]] or "")
    f = str(r[C["F"]] or "")
    if d == "01手机":   return "手机"
    if d == "05电脑":   return "PC"
    if d == "06平板电脑": return "平板"
    if d == "08穿戴":   return "穿戴"
    if d == "07音频":   return "音频"
    if f.startswith("12"): return "HD"
    if "增值" in d:     return "增值"
    return "其他"


def build_day_details(rxs):
    """当日达成明细：RXS 逐行（商品/业务员/金额/毛利/毛利率/成本 + 品类）。"""
    out = []
    for r in rxs:
        amt = num(r[C["M"]])
        gross = num(r[C["N"]])
        out.append({
            "code": str(r[C["A"]]).strip(),
            "emp": str(r[C["P"]]).strip(),
            "product": str(r[C["G"]]).strip(),
            "sku": str(r[C["F"]]).strip(),
            "qty": num(r[C["I"]]),
            "origPrice": num(r[C["K"]]),
            "discPrice": num(r[C["L"]]),
            "amount": amt,
            "profit": gross,
            "cost": num(r[C["S"]]),
            "gpr": (gross / amt) if amt else None,
            "cat": day_cat(r),
        })
    out.sort(key=lambda x: (x["emp"], x["code"]))
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("tsv")
    ap.add_argument("--xlsx", default="/Users/mac/Desktop/李家村销售/李家村8月任务进度.xlsx")
    ap.add_argument("--day", help="当日达成基准日，默认取明细里的最大日期")
    ap.add_argument("-o", "--out", default=os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "data.json"))
    a = ap.parse_args()

    xs = load_tsv(a.tsv)
    if not xs:
        sys.exit("❌ 明细里没有数据行")

    # 当日达成严格锚定「今天」：不把昨天的数据顶上来当今天的日报。
    # data_max=明细里最新出库日；同月且今天≥数据日时 day=今天，否则回退 data_max（跨月/时钟异常）。
    data_max = max(day_of(r) for r in xs)
    today = datetime.date.today()
    base0 = datetime.date.fromisoformat(data_max)
    ref = today if (today.year, today.month) == (base0.year, base0.month) and today >= base0 else base0
    day = a.day or ref.isoformat()
    rxs = [r for r in xs if day_of(r) == day]
    base = datetime.date.fromisoformat(day)
    last_day = calendar.monthrange(base.year, base.month)[1]
    rd = remain_days(ref)

    tasks, lehui, taili, perf_score, lab_day, lab_gap = load_manual(a.xlsx)

    people = {}
    for n in PEOPLE_ORDER:
        p = calc_perf(xs, n, tasks[n])
        q = calc_qcs(xs, n, tasks[n], p, lehui, taili)
        p["绩效"] = perf_score[n]
        people[n] = {
            "performance": p,
            "qcs": q,
            "dailyDone": calc_daily(rxs, n, lab_day),
            "dailyGap": calc_gap(p, q, rd, lab_gap) if n in TASK_PEOPLE else {},
        }

    sp = total_perf(people)
    _ps = [v for v in perf_score.values() if v is not None]      # 合计行是 AVERAGE，忽略空值
    sp["绩效"] = sum(_ps) / len(_ps) if _ps else None
    sq = total_qcs(people, sp)
    store = {
        "performance": sp,
        "qcs": sq,
        "dailyDone": {k: sum(people[n]["dailyDone"].get(k, 0) for n in PEOPLE_ORDER)
                      for k in people[PEOPLE_ORDER[0]]["dailyDone"]},
        "dailyGap": {k: sum(people[n]["dailyGap"].get(k, 0) for n in TASK_PEOPLE)
                     for k in people["邵乐乐"]["dailyGap"]},
    }

    data = {
        "meta": {
            "storeName": "华为李家村万达授权体验店",
            "date": day,
            "dayTitle": f"{base.month:02d}-{base.day:02d}达成",
            "timeProgress": ref.day / last_day,
            "remainDays": rd,
            "refDate": ref.isoformat(),
            # 上账时效：店员销售后才上账，可能滞后，但当天必补齐
            "isToday": base == ref,
            "lagDays": (ref - base).days,
            "todayLabel": f"{ref.month:02d}-{ref.day:02d}",
            "fetchTime": datetime.datetime.now().strftime("%H:%M"),
            "employees": PEOPLE_ORDER,
            "sourceFile": os.path.basename(a.tsv),
            "sourceRows": len(xs),
            "dayRows": len(rxs),
            "generatedAt": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        },
        "store": store,
        "people": people,
        "insights": build_insights(xs, rxs, people, store, ref, ref.day / last_day, rd),
        "details": build_details(xs),
        "dayDetails": build_day_details(rxs),
    }

    with open(a.out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)

    # ---- 体检 ----
    print(f"✅ 复算完成 → {a.out}")
    print(f"   明细 {len(xs)} 行 | 基准日 {day}（当日 {len(rxs)} 行）"
          f" | 时间进度 {base.day/last_day:.1%} | 剩余 {rd} 天")
    g = sp["毛利"]
    print(f"   门店毛利 任务 {g['task']:,.0f} 完成 {g['done']:,.0f} 达成 {(g['rate'] or 0):.1%}")
    print(f"   门店销额 完成 {sp['销额']['done']:,.0f} | 手机 {sp['手机']['done']:.0f} 台"
          f" | 增值 {sq['增值']['done']:,.0f}")
    for n in PEOPLE_ORDER:
        pg = people[n]["performance"]["毛利"]
        tag = " （未分配任务）" if not pg["task"] else ""
        print(f"   - {n:4s} 毛利 {pg['done']:>9,.0f}  当日 {people[n]['dailyDone'].get('毛利',0):>8,.0f}{tag}")


if __name__ == "__main__":
    main()
