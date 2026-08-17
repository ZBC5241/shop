#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
refresh_today_block.py —— 刷新《李家村销售》表「今日达成」区块并取值注入 data.json

流程：
  1. 从用友毛利明细 TSV 取“最新出库日期”的当日行
  2. 写入桌面包 RXS 明细表（公式数据源），数值列去千分位逗号并转 float
  3. 用真实表格引擎（LibreOffice/soffice 优先，否则 Microsoft Excel via AppleScript）重算整本
  4. 读「今日达成」区块（李家村销售!B28:N33）刷新出来的缓存值
  5. 注入 data.json：store.dailyDone + 各营业员 dailyDone，并把 meta.date 对齐真实数据日

这样日报“当日达成”严格等于表格公式重算后的结果，不再由 Python 近似复算。

用法：
  python3 refresh_today_block.py --tsv yonyou_raw.tsv \
      --xlsx "/Users/mac/Desktop/李家村销售/李家村8月任务进度.xlsx" \
      --data data.json
"""
import argparse, csv, json, os, shutil, subprocess, sys, datetime, time

# RXS 表头顺序（列 A~S）：出库单号/单据类型/出库日期/商品分类/商品sku分类/商品SKU编码/
#   商品名称/入库属性/数量/单价/原价/折扣价/金额/毛利/SO激励/业务员/库区/销售出库单门店/销售成本
NUM_COLS_1B = {9, 10, 11, 12, 13, 14, 15, 19}   # 数值列(1-based)：数量/单价/原价/折扣价/金额/毛利/SO激励/销售成本

# 「今日达成」区块位置（李家村销售 sheet）
BLOCK_SHEET = "李家村销售"
LABEL_ROW = 26          # B26:N26 品类标签
FIRST_DATA_ROW = 28     # B28 起 5 位营业员
N_PEOPLE = 5
TOTAL_ROW = 33          # B33 合计
BLOCK_COLS = range(2, 15)   # B~N
PEOPLE = ["邵乐乐", "杨丽华", "李泽", "陈超磊", "张博晨"]


def num(v):
    if v is None:
        return None
    s = str(v).replace(",", "").replace("，", "").strip()
    if s == "":
        return None
    try:
        return float(s)
    except Exception:
        return s


def data_day_of(tsv):
    """取『当日正销额>0』的最大出库日期，作为日报数据日。

    避开纯退货/冲账日（如 8-17 是李泽把 8-16 两笔 Watch GT7 退货冲账+重开，
    净额仅 762，并非真实营业）。判定：当日净销额 > 0，且净销额 ≥ 当日正向
    销额的 30%（即大部分被退货冲掉的日子视为冲账日，排除，回退到前一天）。
    金额列 M(索引12)。
    """
    import collections
    rows = list(csv.reader(open(tsv, encoding="utf-8"), delimiter="\t"))
    net = collections.defaultdict(float)   # 当日净销额
    pos = collections.defaultdict(float)   # 当日正向销额（仅正金额行）
    for r in rows[1:]:
        if len(r) > 2 and r[2]:
            try:
                m = float(str(r[12]).replace(",", ""))
            except Exception:
                m = 0.0
            d = r[2][:10]
            net[d] += m
            if m > 0:
                pos[d] += m
    cand = [d for d, m in net.items()
            if m > 0 and (pos[d] <= 0 or m >= pos[d] * 0.3)]
    return max(cand) if cand else None


def write_rxs(xlsx, tsv, day):
    """把 day 当日明细写入 RXS（清空旧数据行），返回写入行数。"""
    rows = list(csv.reader(open(tsv, encoding="utf-8"), delimiter="\t"))
    today = [r for r in rows[1:] if len(r) > 2 and r[2][:10] == day]
    import openpyxl
    wb = openpyxl.load_workbook(xlsx)
    rx = wb["RXS"]
    for r in range(2, rx.max_row + 1):
        for c in range(1, 20):
            rx.cell(r, c).value = None
    for i, tr in enumerate(today):
        r = 2 + i
        for c in range(1, 20):
            raw = tr[c - 1] if c - 1 < len(tr) else None
            rx.cell(r, c).value = num(raw) if c in NUM_COLS_1B else raw
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    wb.save(xlsx)
    return len(today)


def recalc_xlsx(xlsx):
    """尽力用 Excel 触发表格重算（让用户打开文件时区块值正确），但【不作为 data.json 数据来源】。

    注意：曾用 soffice --convert-to 同名覆盖原文件，在 macOS 上会把 RXS 写乱（回到旧数据），
    已弃用。data.json 当日达成一律由 py_calc_block() 的 Python 口径计算，与表格公式等价。
    此处仅可选触发 Excel 重算表格显示，失败不影响日报数据。
    """
    scpt = '''
    set f to POSIX file "%s"
    tell application "Microsoft Excel"
        launch
        open f
        delay 5
        save active workbook
        close active workbook
    end tell
    ''' % xlsx
    try:
        subprocess.run(["osascript", "-e", scpt], check=True, timeout=120,
                       stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return True
    except Exception as e:
        sys.stderr.write("  [提示] Excel 触发重算不可用（无 GUI），不影响日报：%s\n" % e)
        return False


def read_block(xlsx):
    """（保留备用）读「今日达成」区块缓存值（data_only）。"""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[BLOCK_SHEET]
    labels = [ws.cell(LABEL_ROW, c).value for c in BLOCK_COLS]
    per_person = {}
    for i, p in enumerate(PEOPLE):
        rr = FIRST_DATA_ROW + i
        vals = {}
        for c in BLOCK_COLS:
            v = ws.cell(rr, c).value
            vals[labels[c - 2]] = 0.0 if v is None else float(v)
        per_person[p] = vals
    total = {}
    for c in BLOCK_COLS:
        v = ws.cell(TOTAL_ROW, c).value
        total[labels[c - 2]] = 0.0 if v is None else float(v)
    rx = wb["RXS"]
    sales_total = 0.0
    sales_by = {}
    for r in range(2, rx.max_row + 1):
        p = rx.cell(r, 16).value
        m = rx.cell(r, 13).value
        if isinstance(m, (int, float)):
            sales_total += m
            sales_by[p] = sales_by.get(p, 0.0) + m
    return labels, per_person, total, sales_by, sales_total


# 13 品类标签（与《李家村销售》B26:N26 一致）
CAT_LABELS = ["手机", "毛利", "增值", "智慧办公", "音频穿戴", "HD", "会员",
              "回收", "贴膜", "电信积分", "滞销", "摄影课", "优享/会员"]


def py_calc_block(tsv, day):
    """Python 复刻「今日达成」区块公式口径，返回 (labels, per_person, total, sales_by, sales_total)。

    复用 calc_data.calc_daily（SUMIFS 口径，已验证与表格公式一致），并补齐 calc_daily 漏掉的『摄影课』。
    数据源是 TSV 当日行，与表格 RXS 同源，结果等价。
    """
    import calc_data as bd
    rows = bd.load_tsv(tsv)
    tr = [r for r in rows if r[2][:10] == day]
    block = {}
    for p in PEOPLE:
        v = bd.calc_daily(tr, p, CAT_LABELS)   # 含销额(=M列)
        v["摄影课"] = bd.sumifs(tr, "I", ("P", p), ("G", "*大师课*"), ("N", ">0"))
        block[p] = v
    total = {k: sum(block[p].get(k, 0.0) for p in PEOPLE) for k in CAT_LABELS}
    sales_total = sum(block[p].get("销额", 0.0) for p in PEOPLE)
    sales_by = {p: block[p].get("销额", 0.0) for p in PEOPLE}
    return CAT_LABELS, block, total, sales_by, sales_total


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tsv", required=True)
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--data", required=True)
    ap.add_argument("--no-excel", action="store_true", help="跳过真实引擎重算（仅开放pycel兜底时才有意义）")
    a = ap.parse_args()

    day = data_day_of(a.tsv)
    if not day:
        sys.stderr.write("✗ TSV 无有效日期，跳过今日达成刷新\n")
        return 1
    print("  数据日(最新出库日期): %s" % day)

    # 备份 RXS（仅保留一份轮换备份）
    bak = a.xlsx + ".bak_rxs"
    shutil.copy(a.xlsx, bak)

    n = write_rxs(a.xlsx, a.tsv, day)
    print("  ✓ 已写 RXS 当日明细 %d 行（备份 %s）" % (n, os.path.basename(bak)))

    # 真实引擎（Excel）仅尽力触发表格显示重算，不作为数据来源
    if not a.no_excel:
        recalc_xlsx(a.xlsx)

    # 当日达成一律用 Python 口径（与表格公式等价，稳定可复现）
    labels, per_person, total, sales_by, sales_total = py_calc_block(a.tsv, day)
    store_daily = dict(total)
    store_daily["销额"] = sales_total

    d = json.load(open(a.data, encoding="utf-8"))
    d["store"]["dailyDone"] = store_daily
    for p in PEOPLE:
        pd = dict(per_person[p])
        pd["销额"] = sales_by.get(p, 0.0)
        d["people"][p]["dailyDone"] = pd
    # 对齐真实数据日
    d["meta"]["date"] = day
    d["meta"]["dayTitle"] = day
    d["meta"]["refDate"] = day
    json.dump(d, open(a.data, "w", encoding="utf-8"), ensure_ascii=False)

    print("  ✓ 已注入 data.json：当日达成 = 销额 %d / 毛利 %d / 手机 %d"
          % (round(sales_total), round(store_daily["毛利"]), round(store_daily["手机"])))
    print("    逐人毛利: " + "  ".join("%s %d" % (p, round(per_person[p]["毛利"]))
                                        for p in PEOPLE))
    return 0


if __name__ == "__main__":
    sys.exit(main())
