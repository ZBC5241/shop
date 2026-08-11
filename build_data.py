#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
李家村门店业绩看板 —— 数据抽取器
从《李家村X月任务进度.xlsx》的「李家村销售」表抽取全部指标，生成 data.json

原则：只提取，不计算。表格里是什么数，就输出什么数。
      仅做必要的格式清洗（#DIV/0! -> null，"85%" -> 0.85，去千分位）

用法：
    python build_data.py /path/to/李家村8月任务进度.xlsx [输出.json]
"""
import sys, os, json, datetime
import openpyxl

SHEET = "李家村销售"

# ---------- 表格坐标（1-based 行号，0-based 列索引） ----------
ROW_TIME = 1                       # 时间进度行：B1=日期, I1=进度
PEOPLE_ORDER = ["邵乐乐", "杨丽华", "李泽", "陈超磊", "张博晨"]

# 区块1：业绩考核
P1_ROWS = {"邵乐乐": 4, "杨丽华": 5, "李泽": 6, "陈超磊": 7, "张博晨": 8}
P1_TOTAL_ROW = 9
P1_BLOCKS = [("毛利", 1), ("手机", 5), ("PC", 9), ("平板", 13), ("穿戴", 17),
             ("音频", 21), ("HD", 25), ("智慧办公", 29), ("音频穿戴", 33), ("销额", 37)]
P1_SCORE_COL = 41                  # 绩效

# 区块2：全科生 / 增值
P2_ROWS = {"邵乐乐": 14, "杨丽华": 15, "李泽": 16, "陈超磊": 17, "张博晨": 18}
P2_TOTAL_ROW = 19

# 区块3：当日达成
P3_ROWS = {"邵乐乐": 28, "杨丽华": 29, "李泽": 30, "陈超磊": 31, "张博晨": 32}
P3_TOTAL_ROW = 33
P3_LABEL_ROW = 26                  # 项目名所在行
P3_TITLE_CELL = (25, 2)            # B25 = "08-09达成"

# 区块4：每日缺口（注意：无张博晨）
P4_ROWS = {"邵乐乐": 38, "杨丽华": 39, "李泽": 40, "陈超磊": 41}
P4_TOTAL_ROW = 42
P4_LABEL_ROW = 36


def num(v):
    """清洗成数字或 None。不做任何运算。"""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "" or s.startswith("#"):      # #DIV/0! #N/A #VALUE!
        return None
    s = s.replace(",", "").replace("¥", "").replace("￥", "")
    pct = s.endswith("%")
    if pct:
        s = s[:-1]
    try:
        f = float(s)
    except ValueError:
        return None
    return f / 100 if pct else f


def r4(ws, row, col):
    """读取「任务/完成/缺口/完成率」四连列"""
    return {
        "task": num(ws.cell(row, col + 1).value),
        "done": num(ws.cell(row, col + 2).value),
        "gap":  num(ws.cell(row, col + 3).value),
        "rate": num(ws.cell(row, col + 4).value),
    }


def cell(ws, row, col0):
    """col0 为 0-based 列索引"""
    return num(ws.cell(row, col0 + 1).value)


def read_perf(ws, row):
    d = {name: r4(ws, row, col) for name, col in P1_BLOCKS}
    d["绩效"] = cell(ws, row, P1_SCORE_COL)
    return d


def read_qcs(ws, row):
    return {
        "电信积分":   {"task": cell(ws, row, 1), "done": cell(ws, row, 2),
                     "gap": cell(ws, row, 3), "rate": cell(ws, row, 4)},
        "会员搭售率": {"terminal": cell(ws, row, 5), "care": cell(ws, row, 6),
                     "gap": cell(ws, row, 7), "rate": cell(ws, row, 8)},
        "回收搭售率": {"orders": cell(ws, row, 9), "gap": cell(ws, row, 10),
                     "rate": cell(ws, row, 11)},
        "贴膜率":     {"orders": cell(ws, row, 12), "gap": cell(ws, row, 13),
                     "rate": cell(ws, row, 14)},
        # 摄影课（列15/16）按需求不再提取
        "考核机型":   {"task": cell(ws, row, 17), "gap": cell(ws, row, 18)},
        "乐回收":     {"orders": cell(ws, row, 19), "amount": cell(ws, row, 20),
                     "增值": cell(ws, row, 21)},
        "太力回收":   {"orders": cell(ws, row, 22), "amount": cell(ws, row, 23),
                     "增值": cell(ws, row, 24)},
        "增值":       {"task": cell(ws, row, 25), "done": cell(ws, row, 26),
                     "gap": cell(ws, row, 27), "rate": cell(ws, row, 28)},
        "健康度":     {"coupon": cell(ws, row, 29), "ratio": cell(ws, row, 30),
                     "grossMargin": cell(ws, row, 31), "增值率": cell(ws, row, 32)},
        "星联会员":   {"优享": cell(ws, row, 33), "尊享": cell(ws, row, 34),
                     "合计": cell(ws, row, 35)},
    }


def read_labels(ws, label_row, start=1, end=15):
    out = []
    for c in range(start, end):
        v = ws.cell(label_row, c + 1).value
        out.append(str(v).strip() if v else None)
    return out


def read_qudao(wb):
    """读取「渠道挂账」sheet：时间进度 + 合计行 + 逐人。只提取不计算。"""
    name = "渠道挂账"
    if name not in wb.sheetnames:
        return None
    ws = wb[name]
    # 时间进度：B1=日期, E1=进度率
    raw_date = ws.cell(1, 2).value
    if isinstance(raw_date, (datetime.datetime, datetime.date)):
        date_str = raw_date.strftime("%Y-%m-%d")
    else:
        date_str = str(raw_date).strip()[:10] if raw_date else ""
    tp = num(ws.cell(1, 5).value)

    # 定位表头行（B列=任务 且 C列=完成）
    hrow = None
    for r in range(1, min(ws.max_row, 40) + 1):
        if (str(ws.cell(r, 2).value or "").strip() == "任务"
                and str(ws.cell(r, 3).value or "").strip() == "完成"):
            hrow = r
            break
    if not hrow:
        return None

    people, total = [], None
    for r in range(hrow + 1, ws.max_row + 1):
        nm = ws.cell(r, 1).value
        if nm is None:
            continue
        nm = str(nm).strip()
        if nm == "":
            continue
        if nm == "合计":
            total = {
                "task": num(ws.cell(r, 2).value),
                "done": num(ws.cell(r, 3).value),
                "gap":  num(ws.cell(r, 4).value),
                "rate": num(ws.cell(r, 5).value),
            }
            continue
        people.append({
            "name": nm,
            "task": num(ws.cell(r, 2).value),
            "done": num(ws.cell(r, 3).value),
            "gap":  num(ws.cell(r, 4).value),
            "rate": num(ws.cell(r, 5).value),
        })
    return {"timeDate": date_str, "timeRate": tp, "total": total, "people": people}


# 不再提取的指标（按需求剔除）
DROP_KEYS = {"摄影课"}


def read_flat(ws, row, labels, start=1):
    d = {}
    for i, lab in enumerate(labels):
        if lab and lab not in DROP_KEYS:
            d[lab] = cell(ws, row, start + i)
    return d


def main():
    if len(sys.argv) < 2:
        print("用法: python build_data.py <xlsx路径> [输出json]")
        sys.exit(1)
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "data.json")

    wb = openpyxl.load_workbook(src, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"❌ 找不到工作表「{SHEET}」，现有：{wb.sheetnames}")
        sys.exit(1)
    ws = wb[SHEET]

    # --- 元信息 ---
    raw_date = ws.cell(ROW_TIME, 2).value
    if isinstance(raw_date, (datetime.datetime, datetime.date)):
        date_str = raw_date.strftime("%Y-%m-%d")
    else:
        date_str = str(raw_date).strip()[:10] if raw_date else ""
    tp = num(ws.cell(ROW_TIME, 9).value)

    d3_labels = read_labels(ws, P3_LABEL_ROW)
    d4_labels = read_labels(ws, P4_LABEL_ROW)
    day_title = ws.cell(*P3_TITLE_CELL).value or ""

    data = {
        "meta": {
            "storeName": "华为李家村万达授权体验店",
            "date": date_str,
            "dayTitle": str(day_title).strip(),
            "timeProgress": tp,
            "employees": PEOPLE_ORDER,
            "sourceFile": os.path.basename(src),
            "generatedAt": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        },
        "store": {
            "performance": read_perf(ws, P1_TOTAL_ROW),
            "qcs":         read_qcs(ws, P2_TOTAL_ROW),
            "dailyDone":   read_flat(ws, P3_TOTAL_ROW, d3_labels),
            "dailyGap":    read_flat(ws, P4_TOTAL_ROW, d4_labels),
        },
        "people": {},
    }

    for name in PEOPLE_ORDER:
        data["people"][name] = {
            "performance": read_perf(ws, P1_ROWS[name]),
            "qcs":         read_qcs(ws, P2_ROWS[name]),
            "dailyDone":   read_flat(ws, P3_ROWS[name], d3_labels),
            "dailyGap":    read_flat(ws, P4_ROWS[name], d4_labels) if name in P4_ROWS else {},
        }

    qd = read_qudao(wb)
    if qd:
        data["qudao"] = qd

    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)

    # --- 体检报告 ---
    print(f"✅ 已生成 {out}")
    print(f"   数据日期 {date_str} | 时间进度 {tp:.1%}" if tp else f"   数据日期 {date_str}")
    g = data["store"]["performance"]["毛利"]
    print(f"   门店毛利 任务{g['task']} 完成{g['done']} 达成{(g['rate'] or 0):.1%}")
    for n in PEOPLE_ORDER:
        pg = data["people"][n]["performance"]["毛利"]
        tag = "（未分配任务）" if pg["task"] in (None, 0) else ""
        print(f"   - {n}: 完成 {pg['done']}  {tag}")


if __name__ == "__main__":
    main()
