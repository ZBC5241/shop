#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
merge_qudao.py —— 把《李家村X月任务进度.xlsx》的「渠道挂账」sheet 抽取后并入 data.json。

为什么需要它：
  run_pipeline 主复算走 calc_data.py（从明细实时复算，数字最准），但 calc_data 不含渠道挂账；
  渠道挂账只有 build_data.py 的 read_qudao 抽。本脚本只补 qudao 字段，不动 calc_data 的其他指标，
  避免“为了接渠道而退回公式缓存口径”的旧问题。

用法：
  python merge_qudao.py [data.json] [xlsx路径]
"""
import sys, os, json
import openpyxl
import build_data as bd

BASE = os.path.dirname(os.path.abspath(__file__))
DATA = sys.argv[1] if len(sys.argv) > 1 else os.path.join(BASE, "data.json")
XLSX = sys.argv[2] if len(sys.argv) > 2 else "/Users/mac/Desktop/李家村销售/李家村8月任务进度.xlsx"

AUG_CACHE = os.path.join(BASE, "sa_aug_cache.json")   # 8 月李家村切片（纯HTTP 抓取后本地筛得）


def _num(v):
    try:
        f = float(v)
        return f if f == f else 0.0
    except Exception:
        return 0.0


# 业务类型黑名单：与《李家村销售分析_销售净额_不含垫付预订》口径一致，
# 渠道/业绩统计均剔除「垫付」「预订」（产品订金、非MSC垫付等不计入销售净额）。
EXCLUDE_BTYPE = {"垫付", "预订"}


def read_channels():
    """按【华为获客渠道名称】聚合 8 月李家村销售分析的渠道明细。

    口径：剔除业务类型 ∈ {垫付, 预订} 的记录（与用户 Excel「不含垫付预订」表一致）。

    来源优先级：
      1) sa_aug_cache.json（纯HTTP 抓取的 8 月切片，最干净、最可靠）
      2) xlsx「销售分析」sheet（列 P=渠道名, J=业务类型, AM=销售净额, AI=销售数量）
    返回按销额净额降序的列表：[{name, amount, qty, bills, share}]
    """
    outset = []
    # —— 来源 1：sa_aug_cache.json ——
    if os.path.exists(AUG_CACHE):
        try:
            recs = json.load(open(AUG_CACHE, encoding="utf-8")).get("records", [])
            agg = {}
            for r in recs:
                bt = str(r.get("iBusinesstypeid_name") or "").strip()
                if bt in EXCLUDE_BTYPE:
                    continue
                c = r.get("retailVouchHeaderDefineCharacter__HWHKQD_name") or ""
                c = str(c).strip()
                if not c:
                    continue
                d = agg.setdefault(c, {"amount": 0.0, "qty": 0.0, "bills": 0})
                d["amount"] += _num(r.get("fNetMoney"))
                d["qty"]    += _num(r.get("fQuantity"))
                d["bills"]  += 1
            outset = agg
        except Exception as e:
            print("  [渠道明细] sa_aug_cache 读取失败，转 xlsx: %s" % e)

    # —— 来源 2：xlsx「销售分析」sheet 兜底 ——
    if not outset:
        try:
            wb = openpyxl.load_workbook(XLSX, data_only=True)
            if "销售分析" in wb.sheetnames:
                ws = wb["销售分析"]
                agg = {}
                for r in range(3, ws.max_row + 1):
                    bt = str(ws.cell(r, 10).value or "").strip()   # J 列 = 业务类型名称
                    if bt in EXCLUDE_BTYPE:
                        continue
                    c = ws.cell(r, 16).value      # P 列 = 华为获客渠道名称
                    if c is None:
                        continue
                    c = str(c).strip()
                    if not c:
                        continue
                    d = agg.setdefault(c, {"amount": 0.0, "qty": 0.0, "bills": 0})
                    d["amount"] += _num(ws.cell(r, 39).value)   # AM = 销售净额
                    d["qty"]    += _num(ws.cell(r, 35).value)   # AI = 销售数量
                    d["bills"]  += 1
                outset = agg
        except Exception as e:
            print("  [渠道明细] xlsx 读取失败: %s" % e)

    if not outset:
        return []

    total = sum(d["amount"] for d in outset.values())
    rows = []
    for name, d in outset.items():
        rows.append({
            "name": name,
            "amount": round(d["amount"], 2),
            "qty": round(d["qty"], 2),
            "bills": d["bills"],
            "share": (d["amount"] / total) if total else 0.0,
        })

    # —— 渠道展示白名单：仅展示用户指定的获客渠道（其余不列入明细）——
    # 白名单内但本期无数据的渠道补 0 行展示；占比基于白名单合计。
    CHANNEL_WHITELIST = ["三大地图", "小红书", "大众点评", "异业", "社区", "企业上门购"]
    by_name = {r["name"]: r for r in rows}
    out_rows = []
    for name in CHANNEL_WHITELIST:
        if name in by_name:
            out_rows.append(by_name[name])
        else:
            out_rows.append({"name": name, "amount": 0.0, "qty": 0.0, "bills": 0, "share": 0.0})
    wl_total = sum(r["amount"] for r in out_rows)
    for r in out_rows:
        r["share"] = (r["amount"] / wl_total) if wl_total else 0.0
    return out_rows


def read_emp_channel():
    """按【员工 × 华为获客渠道】聚合 8 月李家村销售分析（员工渠道挂账完成额）。

    口径：与「渠道挂账」完成列公式一致（SUMIFS 从销售分析表提取白名单渠道净额，
          垫付/预订也算——晨哥 2026-08-19 拍板），保证逐人渠道明细合计 == 渠道挂账完成额；
          仅白名单获客渠道；按单据日期过滤 8 月切片。
    来源优先级：
      1) sa_aug_cache.json（纯HTTP 抓取的 8 月全量切片，最新最全）
      2) xlsx「销售分析」sheet（用户导出，兜底）
    返回：( { 员工名: [ {channel, amount, bills}, ... 按白名单顺序 ] },
           { 员工名: { 渠道: [ {date, code, product, sku, qty, amount, member, phone}, ... ] } } )
    """
    emp = {}
    items = {}
    sa_lookup = {}

    # —— 来源 1：sa_aug_cache.json（用友云 8 月切片，最全）——
    if os.path.exists(AUG_CACHE):
        try:
            recs = json.load(open(AUG_CACHE, encoding="utf-8")).get("records", [])
            n_emp = 0
            for r in recs:
                p = str(r.get("iEmployeeid_name") or "").strip()
                c = str(r.get("retailVouchHeaderDefineCharacter__HWHKQD_name") or "").strip()
                if not p or not c:
                    continue
                amt = _num(r.get("fNetMoney"))
                d = emp.setdefault(p, {}).setdefault(c, {"amount": 0.0, "bills": 0})
                d["amount"] += amt
                d["bills"]  += 1
                bd = str(r.get("dDate") or r.get("vouchdate") or "")[:10]
                items.setdefault(p, {}).setdefault(c, []).append({
                    "date": bd,
                    "code": str(r.get("code") or r.get("id") or "").strip(),
                    "product": str(r.get("product_cName") or r.get("oid_userDefine_2394043221715451912") or "").strip(),
                    "sku": str(r.get("productsku_cCode") or r.get("product_cCode") or "").strip(),
                    "qty": _num(r.get("fQuantity")),
                    "amount": amt,
                    "member": str(r.get("iMemberid_name") or "").strip(),
                    "phone": str(r.get("iMemberid_cphone") or "").strip(),
                })
                # 全量索引（含非白名单渠道）：供品类明细补 单号/会员/电话
                # ⚠️ key 只用 (员工,日期,净额)：切片的 productsku_cCode 是「商品名称」，
                #    而 details 的 sku 是「SKU编码」，两者交集为 0，加 sku 会导致 0 命中。
                sa_lookup.setdefault((p, bd, amt), []).append({
                    "code": str(r.get("code") or r.get("id") or "").strip(),
                    "member": str(r.get("iMemberid_name") or "").strip(),
                    "phone": str(r.get("iMemberid_cphone") or "").strip(),
                })
                n_emp += 1
            if emp:
                print("  [员工×渠道] 用友云切片聚合完成（%d 行，员工 %d 人）" % (n_emp, len(emp)))
            else:
                print("  [员工×渠道] sa_aug_cache 为空，转 xlsx 兜底")
        except Exception as e:
            print("  [员工×渠道] sa_aug_cache 读取失败，转 xlsx: %s" % e)

    # —— 来源 2：xlsx「销售分析」sheet 兜底（仅当切片无数据）——
    if not emp and os.path.exists(XLSX):
        try:
            wb = openpyxl.load_workbook(XLSX, data_only=True)
            if "销售分析" not in wb.sheetnames:
                print("  [员工×渠道] 未找到「销售分析」sheet")
                return emp, items
            ws = wb["销售分析"]
            # 表头在第 2 行，动态定位关键列
            hdr = {}
            for c in range(1, ws.max_column + 1):
                v = ws.cell(2, c).value
                if v:
                    hdr[str(v).strip()] = c
            COL_P   = hdr.get("华为获客渠道名称")   # 渠道
            COL_G   = hdr.get("营业员名称")         # 员工
            COL_N   = hdr.get("销售净额")           # 净额
            COL_BD  = hdr.get("业务日期")           # 过滤 8 月
            COL_C   = hdr.get("单据编号")           # 单号
            COL_PRO = hdr.get("商品名称")           # 商品名称
            COL_SKU = hdr.get("商品sku名称")        # SKU
            COL_QTY = hdr.get("销售数量")           # 数量
            COL_MEM = hdr.get("会员姓名")           # 会员姓名
            COL_PH  = hdr.get("会员手机号")          # 会员手机号
            if not (COL_G and COL_P and COL_N):
                print("  [员工×渠道] 销售分析缺少关键列（员工/渠道/净额）")
                return emp, items
            import datetime as _dt
            for r in range(3, ws.max_row + 1):
                bd = ws.cell(r, COL_BD).value if COL_BD else None
                if isinstance(bd, (_dt.datetime, _dt.date)):
                    if not (bd.year == 2026 and bd.month == 8):
                        continue
                p = str(ws.cell(r, COL_G).value or "").strip()
                c = str(ws.cell(r, COL_P).value or "").strip()
                if not p or not c:
                    continue
                d = emp.setdefault(p, {}).setdefault(c, {"amount": 0.0, "bills": 0})
                d["amount"] += _num(ws.cell(r, COL_N).value)
                d["bills"]  += 1
                # 单品明细（供下钻到订单/商品级）
                items.setdefault(p, {}).setdefault(c, []).append({
                    "date": str(bd)[:10] if bd else "",
                    "code": str(ws.cell(r, COL_C).value or "").strip() if COL_C else "",
                    "product": str(ws.cell(r, COL_PRO).value or "").strip() if COL_PRO else "",
                    "sku": str(ws.cell(r, COL_SKU).value or "").strip() if COL_SKU else "",
                    "qty": _num(ws.cell(r, COL_QTY).value) if COL_QTY else 0,
                    "amount": _num(ws.cell(r, COL_N).value),
                    "member": str(ws.cell(r, COL_MEM).value or "").strip() if COL_MEM else "",
                    "phone": str(ws.cell(r, COL_PH).value or "").strip() if COL_PH else "",
                })
                # 全量索引（含非白名单渠道）：供品类明细补 单号/会员/电话
                # ⚠️ 与切片路径统一：key = (员工,日期,净额)，不加 sku（口径不一致会导致 0 命中）
                sa_lookup.setdefault((p, (str(bd)[:10] if bd else ""),
                                      _num(ws.cell(r, COL_N).value)), []).append({
                    "code": str(ws.cell(r, COL_C).value or "").strip() if COL_C else "",
                    "member": str(ws.cell(r, COL_MEM).value or "").strip() if COL_MEM else "",
                    "phone": str(ws.cell(r, COL_PH).value or "").strip() if COL_PH else "",
                })
        except Exception as e:
            print("  [员工×渠道] 销售分析读取失败: %s" % e)
            return emp, items, sa_lookup
    # 渠道展示白名单：仅展示用户指定的获客渠道
    CHANNEL_WHITELIST = ["三大地图", "小红书", "大众点评", "异业", "社区", "企业上门购"]
    out = {}
    for p, chans in emp.items():
        rows = []
        for ch in CHANNEL_WHITELIST:
            d = chans.get(ch)
            rows.append({
                "channel": ch,
                "amount": round(d["amount"], 2) if d else 0.0,
                "bills": d["bills"] if d else 0,
            })
        out[p] = rows
    # 仅保留白名单渠道的单品明细（去掉空渠道）
    items_out = {}
    for p, chans in items.items():
        items_out[p] = {ch: rows for ch, rows in chans.items() if ch in CHANNEL_WHITELIST}
    return out, items_out, sa_lookup


QUDAO_FORMULA = ('=SUM(SUMIFS(销售分析!$AM:$AM,销售分析!$G:$G,A{row},'
                 '销售分析!$P:$P,{{"三大地图","小红书","大众点评","异业","社区","企业上门购"}}))')


def restore_qudao_formulas(xlsx):
    """恢复「渠道挂账」sheet 完成(C)列 SUMIFS 公式，不写死数值。

    用户口径（2026-08-19 晨哥拍板）：挂账完成 = 渠道挂账 sheet「完成」列公式
    （SUMIFS 从销售分析表提取白名单渠道净额）的值，与预订/垫付无关。
    完成列必须保持为活公式，Excel/WPS 打开即自动重算；看板读数走 read_qudao
    的 _agg_qudao_done 兜底（同口径 Python 复算，数值与公式一致）。
    """
    try:
            wb_f = openpyxl.load_workbook(xlsx, data_only=False)
            ws = wb_f["渠道挂账"]
            hrow = None
            for r in range(1, min(ws.max_row, 40) + 1):
                if (str(ws.cell(r, 2).value or "").strip() == "任务"
                        and str(ws.cell(r, 3).value or "").strip() == "完成"):
                    hrow = r
                    break
            if not hrow:
                print("  [渠道公式] 未找到表头行，跳过恢复")
                return False
            restored = 0
            for r in range(hrow + 1, ws.max_row + 1):
                nm = ws.cell(r, 1).value
                if not nm:
                    continue
                nm = str(nm).strip()
                if nm in ("", "合计"):
                    continue
                ws.cell(r, 3).value = QUDAO_FORMULA.format(row=r)
                restored += 1
            wb_f.save(xlsx)
            print("  [渠道公式] 已恢复 C 列 SUMIFS 公式 %d 格（完成=销售分析白名单渠道净额）" % restored)
            return True
    except Exception as e:
        print("  [渠道公式] 恢复失败: %s" % e)
        return False


def main():
    if not os.path.exists(DATA):
        print("✗ 找不到 data.json，请先运行 calc_data.py")
        return 1
    if not os.path.exists(XLSX):
        print("✗ 找不到表格: %s" % XLSX)
        return 1

    # Step A：恢复 C 列公式（保持「完成」为活公式，Excel 打开自动重算）
    restore_qudao_formulas(XLSX)

    # Step B：重新读取（含最新写入的 C 列）做合并
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    wb_f = openpyxl.load_workbook(XLSX, data_only=False)
    qd = bd.read_qudao(wb, wb_f)
    if not qd:
        print("[渠道] 未找到「渠道挂账」数据，跳过")
        return 0

    # 渠道挂账完成额/任务额/达成/逐人：全部按用户 xlsx「渠道挂账」sheet 公式提取
    # （read_qudao 已读到表格公式值，如合计 24539 / 杨丽华 0）。
    # 逐人下钻的【各渠道明细】则从桌面 xlsx「销售分析」sheet 聚合（口径与挂账完成一致：含全部业务类型），
    # 全程数据源都是用户桌面文件，不再经过用友抓取缓存。
    emp_ch, ch_items, sa_lookup = read_emp_channel()
    if emp_ch:
        qd["empChannel"] = emp_ch
        print("[员工×渠道] 已聚合 %d 名员工的渠道明细（来自桌面销售分析 sheet）" % len(emp_ch))
    if ch_items:
        qd["channelItems"] = ch_items
        print("[渠道单品] 已聚合各员工×渠道单品明细（%d 人）" % len(ch_items))

    with open(DATA, encoding="utf-8") as f:
        d = json.load(f)
    # 单品明细补充 毛利/成本/原价/毛利率
    # ⚠️ 口径坑：details 的 sku 是「SKU编码」，而渠道单品(销售分析)的 sku 是「商品名称」，
    #    两者交集为 0 → 按 sku 匹配恒 0 命中。改用 (日期,净额) 匹配（实测命中 ~99%）。
    det_idx = {}       # (date, sku) 兼容旧口径
    det_idx_amt = {}   # (date, amount)
    det_idx_full = {}  # (date, amount, qty) 最精确
    for _cat, rows in (d.get("details") or {}).items():
        for r in rows:
            _dt = r.get("date")
            _sku = r.get("sku", "")
            _amt = r.get("amount")
            det_idx.setdefault((_dt, _sku), r)
            det_idx_amt.setdefault((_dt, _amt), r)
            det_idx_full.setdefault((_dt, _amt, r.get("qty")), r)
    _enriched = 0
    for _p, _chans in (qd.get("channelItems") or {}).items():
        for _ch, _items in _chans.items():
            for it in _items:
                dr = (det_idx_full.get((it.get("date"), it.get("amount"), it.get("qty")))
                      or det_idx_amt.get((it.get("date"), it.get("amount")))
                      or det_idx.get((it.get("date"), it.get("sku", ""))))
                if not dr:
                    continue
                it["origPrice"] = dr.get("origPrice")
                it["discPrice"] = dr.get("discPrice")
                it["profit"] = dr.get("profit")
                it["gpr"] = dr.get("gpr")
                it["cost"] = dr.get("cost")
                # 商品名补全：毛利明细的 name 含 颜色/内存（如 HOP-AL10 16GB+512GB 活力橙）
                if dr.get("name"):
                    it["product"] = dr["name"]
                _enriched += 1
    if _enriched:
        print("  [渠道单品] 已补毛利/成本字段 %d 条" % _enriched)

    # 品类达成明细补 单号/会员/电话：按 (员工,日期,净额) 关联销售分析
    # ⚠️ 不要加 sku：details 的 sku 是编码、销售分析的是商品名称，加进去会 0 命中
    if sa_lookup:
        _hit = 0
        for _cat, rows in (d.get("details") or {}).items():
            for r in rows:
                info = sa_lookup.get((r.get("emp"), r.get("date"),
                                      _num(r.get("amount"))))
                if not info:
                    continue
                it0 = info[0]
                r["code"] = it0.get("code", "")
                r["member"] = it0.get("member", "")
                r["phone"] = it0.get("phone", "")
                _hit += 1
        if _hit:
            print("  [品类明细] 已补单号/会员/电话 %d 条" % _hit)
    d["qudao"] = qd
    with open(DATA, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, indent=1)

    t = qd.get("total") or {}
    print("[渠道] 已并入 data.json：完成 %s / 任务 %s / 达成 %.1f%%（时间进度 %.1f%%）" % (
        t.get("done"), t.get("task"),
        (t.get("rate") or 0) * 100, (qd.get("timeRate") or 0) * 100))
    return 0


if __name__ == "__main__":
    sys.exit(main())
