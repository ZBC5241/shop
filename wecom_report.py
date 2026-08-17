#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
wecom_report.py —— 生成「今日达成日报」，推送到企业微信群机器人。

读取 data.json（看板复算结果）+ yonyou_raw.tsv（算每人单量），
按「当日任务」算各板块达成率与缺口，拆解到人，
同时推送：
  1) Markdown V2（小红书风 · 每指标一行）
  2) HTML 高清海报图（poster_v3 风格 · 真实数据）

用法:
  python3 wecom_report.py            # 生成并推送（定时任务用）
  python3 wecom_report.py --dry      # 仅打印预览，不推送
"""
import json
import csv
import os
import sys
import base64
import calendar
import subprocess
import urllib.request

BASE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(BASE, "data.json")
TSV = os.path.join(BASE, "yonyou_raw.tsv")
CONF = os.path.join(BASE, ".wecom_webhook")
POSTER_PY = os.path.join(BASE, "poster_build.py")
POSTER_PNG = os.path.join(BASE, "poster_today.png")
PYTHON = sys.executable

# 李家村月度任务进度表（含「渠道挂账」sheet）
TARGET_XLSX = "/Users/mac/Desktop/李家村销售/李家村8月任务进度.xlsx"

# 增值当日任务（晨哥定：周内 1600/天）。其余板块按「全月任务 ÷ 当月工作日」均摊。
DAILY_VALUEADDED = 1600


def load_webhook():
    if os.environ.get("WECOM_WEBHOOK"):
        return os.environ["WECOM_WEBHOOK"].strip()
    if os.path.exists(CONF):
        v = open(CONF, encoding="utf-8").read().strip()
        if v:
            return v
    return None


def num(s):
    s = "" if s is None else str(s)
    s = s.replace(",", "").replace("%", "").strip()
    try:
        return float(s)
    except Exception:
        return 0.0


def pct(x):
    try:
        return "{:.0f}%".format(float(x) * 100)
    except Exception:
        return str(x)


def pct1(x):
    try:
        return "{:.1f}%".format(float(x) * 100)
    except Exception:
        return str(x)


def working_days(year, month):
    """当月工作日（周一~周五）天数。"""
    ndays = calendar.monthrange(year, month)[1]
    return sum(1 for d in range(1, ndays + 1) if calendar.weekday(year, month, d) < 5)


def today_rows(tsv, today):
    """返回今日所有出库行（list of dict）。"""
    try:
        rows = list(csv.reader(open(tsv, encoding="utf-8-sig"), delimiter="\t"))
    except Exception:
        return []
    if not rows:
        return []
    hdr = rows[0]
    out = []
    for r in rows[1:]:
        if not r or not r[0].strip():
            continue
        rec = {hdr[i]: (r[i] if i < len(r) else "") for i in range(len(hdr))}
        if rec.get("出库日期", "").strip()[:10] == today:
            out.append(rec)
    return out


def read_qudao():
    """读取「渠道挂账」：
    - 任务额(B列)、完成额(C列)、时间进度(B1=TODAY()) 均直接取自 xlsx「渠道挂账」sheet
      （C 列「完成」即「最先拉取的数据」，不再实时复算销售分析）。
    - 仅当 C 列全为空时才回退复算一次，避免空表报错。"""
    import openpyxl, datetime, calendar
    import build_data as bd

    SRC = TARGET_XLSX
    try:
        wb = openpyxl.load_workbook(SRC, data_only=True)
    except Exception:
        return None
    if "渠道挂账" not in wb.sheetnames:
        return None
    ws = wb["渠道挂账"]

    # 时间进度（按今天算，等价于表格 B1=TODAY()）
    today = datetime.date.today()
    time_date = today.strftime("%Y-%m-%d")
    time_rate = "{:.1f}%".format(today.day / calendar.monthrange(today.year, today.month)[1] * 100)

    # 定位表头行（B=任务 且 C=完成）
    hrow = None
    for r in range(1, min(ws.max_row, 40) + 1):
        if str(ws.cell(r, 2).value or "").strip() == "任务" and str(ws.cell(r, 3).value or "").strip() == "完成":
            hrow = r
            break
    if not hrow:
        return None

    # 完成额：直接取 sheet C 列「最先拉取」的值；整列空才回退复算
    names, tasks, done_map = [], {}, {}
    for r in range(hrow + 1, ws.max_row + 1):
        nm = ws.cell(r, 1).value
        if nm is None:
            continue
        nm = str(nm).strip()
        if not nm or nm == "合计":
            continue
        names.append(nm)
        tasks[nm] = num(ws.cell(r, 2).value)          # 任务额：本 sheet B 列
        done_map[nm] = num(ws.cell(r, 3).value)       # 完成额：本 sheet C 列

    use_sheet = any(v for v in done_map.values())
    agg = bd._agg_qudao_done(wb) if not use_sheet else {}

    people, tot_task, tot_done = [], 0.0, 0.0
    for n in names:
        if use_sheet:
            done = done_map.get(n, 0.0)               # 完成额：取 sheet C 列「最先拉取」的值
        else:
            done = round(agg.get(n, 0.0), 2)          # 回退：复算 C 列公式
        task = tasks.get(n, 0.0)
        tot_task += task
        tot_done += done
        rate = "{:.1f}%".format(done / task * 100) if task else "0.0%"
        people.append({"name": n, "task": task, "done": done,
                       "gap": task - done, "rate": rate})
    tot_rate = tot_done / tot_task * 100 if tot_task else 0.0
    return {"people": people, "total_task": tot_task, "total_done": tot_done,
            "total_gap": tot_task - tot_done, "total_rate": tot_rate,
            "time_date": time_date, "time_rate": time_rate}


def wan(x):
    """渠道挂账精准值：整数原样显示，不缩写「万」。"""
    x = float(x or 0)
    return "{:.0f}".format(x)


def build_markdown_v2(d):
    """生成「Markdown V2」小红书风日报。无 ¥、无逗号、每指标一行。"""
    meta = d.get("meta", {})
    today = meta.get("date", "")
    dd = d.get("store", {}).get("dailyDone", {})
    perf = d.get("store", {}).get("performance", {})
    emp = meta.get("employees", [])

    rows = today_rows(TSV, today)
    cnt = {}
    for r in rows:
        e = (r.get("业务员", "") or "").strip() or "(未分配)"
        cnt[e] = cnt.get(e, 0) + 1

    amt = dd.get("销额", 0) or 0
    gp = dd.get("毛利", 0) or 0
    va = dd.get("增值", 0) or 0
    phone = dd.get("手机", 0) or 0
    gm = (gp / amt) if amt else 0

    y, m = int(today[:4]), int(today[5:7])
    wd = working_days(y, m)

    def dtask(key):
        if key == "增值":
            return float(DAILY_VALUEADDED)
        t = (perf.get(key, {}) or {}).get("task", 0) or 0
        return t / wd if wd else 0

    def rate(done, task):
        return (done / task) if task else 0

    # 标题日期永远跟随真实数据日（meta.date），避免与内容错位
    label = today[5:] if len(today) == 10 else (meta.get("todayLabel") or today)
    ftime = meta.get("fetchTime", "")

    L = []
    # 无人上账提醒（由 gen_daily_html.py 写入 .daily_alert_msg，营业时段才会写）
    _alert = ""
    try:
        with open(os.path.join(BASE, ".daily_alert_msg"), encoding="utf-8") as f:
            _alert = f.read().strip()
    except Exception:
        _alert = ""
    if _alert:
        L.append("⏰ {}".format(_alert))
        L.append("")
    L.append("# 📊 李家村门店今日达成日报 · {}".format(label))
    L.append('<font color="comment">更新 {} ｜ 以已上账为准</font>'.format(ftime))
    L.append("")

    L.append("**一、今日核心**")
    L.append("销额 {} ｜ 毛利 {} ｜ 增值 {} ｜ 手机 {}台".format(
        "{:.0f}".format(amt), "{:.0f}".format(gp), "{:.0f}".format(va), "{:.0f}".format(phone)))
    L.append("")

    L.append("**二、板块任务达成**")
    blocks = [
        ("销额", amt, dtask("销额")),
        ("毛利", gp, dtask("毛利")),
        ("手机", phone, dtask("手机")),
        ("增值", va, dtask("增值")),
    ]
    for name, done, task in blocks:
        r = rate(done, task)
        if name == "手机":
            L.append("· {}　{:.0f}台 / {:.0f}台 → {}".format(name, done, task, pct(r)))
        else:
            L.append("· {}　{:.0f} / {:.0f} → {}".format(name, done, task, pct(r)))
    L.append("")

    # 三、当日品类达成（按《李家村销售》今日达成区块 13 品类字段，全量展示）
    L.append("**三、当日品类达成 · 按《李家村销售》公式**")
    _cat_order = ["手机", "毛利", "增值", "智慧办公", "音频穿戴", "HD", "会员",
                  "回收", "贴膜", "电信积分", "滞销", "摄影课", "优享/会员"]
    _cat_parts = []
    for k in _cat_order:
        v = dd.get(k, 0) or 0
        if k == "手机":
            _cat_parts.append("{} {}台".format(k, "{:.0f}".format(v)))
        else:
            _cat_parts.append("{} {}".format(k, "{:.0f}".format(v)))
    L.append("　".join(_cat_parts))
    L.append("")

    L.append("**四、人员战况**")
    # 开单在前
    for e in sorted(emp, key=lambda x: -(d.get("people", {}).get(x, {}).get("dailyDone", {}).get("销额", 0) or 0)):
        pd = d.get("people", {}).get(e, {}).get("dailyDone", {})
        e_amt = pd.get("销额", 0) or 0
        e_gp = pd.get("毛利", 0) or 0
        n = cnt.get(e, 0)
        if e_amt == 0:
            L.append("{}　⚠️ 挂零".format(e))
        else:
            L.append("{}　✅ 销额 {} ／ 毛利 {} ／ {}单".format(e, "{:.0f}".format(e_amt), "{:.0f}".format(e_gp), n))
    L.append("")

    # 五、渠道挂账（来自任务进度表独立 sheet）
    qd = read_qudao()
    if qd:

        def rate_str(s):
            try:
                f = float(str(s).replace("%", ""))
                if 0 < f < 1:
                    f *= 100
                return "{:.1f}%".format(f)
            except Exception:
                return str(s)

        L.append("**五、渠道挂账**")
        lag = ""
        if qd["time_rate"]:
            try:
                tr = float(str(qd["time_rate"]).replace("%", ""))
                if 0 < tr < 1:
                    tr *= 100
                if qd["total_rate"] < tr:
                    lag = " ⚠️ 落后"
            except Exception:
                pass
            L.append("完成 {} ／ 任务 {} ／ 达成 {:.1f}%　（时间进度 {}）{}".format(
                wan(qd["total_done"]), wan(qd["total_task"]), qd["total_rate"],
                rate_str(qd["time_rate"]), lag))
        else:
            L.append("完成 {} ／ 任务 {} ／ 达成 {:.1f}%".format(
                wan(qd["total_done"]), wan(qd["total_task"]), qd["total_rate"]))
        _peo = []
        for p in qd["people"]:
            if p["task"] == 0 and p["done"] == 0:
                _peo.append("{} 无任务".format(p["name"]))
            elif p["done"] == 0:
                _peo.append("{} 挂零".format(p["name"]))
            elif p["task"] == 0:
                _peo.append("{} {}（无任务）".format(p["name"], wan(p["done"])))
            else:
                _peo.append("{} {}（{}）".format(p["name"], wan(p["done"]), rate_str(p["rate"])))
        L.append("　".join(_peo))
        L.append("")

    L.append("**六、复盘**")
    L.append("增值电信双低；毛利率 {} 处低位".format(pct1(gm)))

    return "\n".join(L)


def build_image(dry=False):
    """调用 poster_build.py 生成高清海报，返回 PNG bytes；失败返回 None。"""
    if dry:
        return None
    try:
        subprocess.run([PYTHON, POSTER_PY, "--out", POSTER_PNG], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        with open(POSTER_PNG, "rb") as f:
            return f.read()
    except Exception as e:
        print("⚠️ 海报生成失败:", e, file=sys.stderr)
        return None


def send_markdown(webhook, content):
    payload = json.dumps(
        {"msgtype": "markdown", "markdown": {"content": content}},
        ensure_ascii=False,
    ).encode("utf-8")
    req = urllib.request.Request(webhook, data=payload, headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return resp.read().decode("utf-8")
    except Exception as e:
        return "ERR {}".format(e)


def send_image(webhook, png_bytes):
    md5 = __import__("hashlib").md5(png_bytes).hexdigest()
    b64 = base64.b64encode(png_bytes).decode("utf-8")
    payload = json.dumps(
        {"msgtype": "image", "image": {"base64": b64, "md5": md5}},
        ensure_ascii=False,
    ).encode("utf-8")
    req = urllib.request.Request(webhook, data=payload, headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            return resp.read().decode("utf-8")
    except Exception as e:
        return "ERR {}".format(e)


if __name__ == "__main__":
    dry = "--dry" in sys.argv
    wh = load_webhook()
    if not wh and not dry:
        print("⚠️ 未配置企业微信 webhook（.wecom_webhook 或 WECOM_WEBHOOK），跳过推送。")
        sys.exit(0)

    d = json.load(open(DATA, encoding="utf-8"))
    content = build_markdown_v2(d)

    print("--- Markdown V2 预览 ---")
    print(content)
    print("------------------------")

    if dry:
        print("（--dry 模式，未生成海报、未推送）")
        sys.exit(0)

    print("推送 Markdown ...")
    print(send_markdown(wh, content))
