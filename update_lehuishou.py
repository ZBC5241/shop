#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
李家村销售看板 · 乐回收更新工具
==================================

**职责**：
    把"乐回收"当日明细写入《李家村销售》sheet 的固定单元格：
        T14:T18  ← 单量（成交量/总单量）
        U14:U18  ← 增值（公司净利）

    写入完成后，由表内 SUM 公式 T19/U19 自动汇总，cal_data.py 复算时直接读。
    **xlsx 文件即为持久化**，不需要额外的 JSON / 数据库。

**写入单元格映射（铁律·不可改）**：
        行14 → 邵乐乐      行16 → 李泽
        行15 → 杨丽华      行17 → 陈超磊
                       行18 → 张博晨

**用法**：
    # JSON 字符串（推荐）
    python update_lehuishou.py --xlsx 路径 \\
        --data '{"邵乐乐":[8,138],"李泽":[7,451],"陈超磊":[6,1197],"杨丽华":[5,1067]}'

    # 交互式（直接输入一行一人）
    python update_lehuishou.py --xlsx 路径
    > 邵乐乐 8 138
    > 李泽 7 451
    > ...
    > .

    # 只更新单人
    python update_lehuishou.py --xlsx 路径 --邵乐乐 8 138
"""
import argparse, json, os, sys, datetime
import openpyxl

LEHUI_SHEET = "李家村销售"
LEHUI_ROWS = {
    "邵乐乐": 14,
    "杨丽华": 15,
    "李泽":  16,
    "陈超磊": 17,
    "张博晨": 18,
}
# T=单量列(20), U=增值列(21); T19/U19 是 SUM 公式, 不要碰
LEHUI_COL_T = 20
LEHUI_COL_U = 21


def _num(v):
    """数字清洗：空/None → 0；带逗号 → 去逗号；其他 → float。"""
    if v is None or (isinstance(v, str) and not v.strip()):
        return 0.0
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def update_lehuishou(xlsx, data, verify=True):
    """把 {姓名:(单量, 增值)} 写到 T/U 列。

    返回更新后的 {姓名: {orders, amount}}，便于调用方核对 / 看板展示。
    """
    if not os.path.exists(xlsx):
        sys.exit(f"❌ 找不到表: {xlsx}")
    wb = openpyxl.load_workbook(xlsx)
    if LEHUI_SHEET not in wb.sheetnames:
        sys.exit(f"❌ 表里没有 [{LEHUI_SHEET}] sheet: {xlsx}")
    ws = wb[LEHUI_SHEET]

    # 写入前快照（万一用户想撤回）
    snapshot = {}
    for n, r in LEHUI_ROWS.items():
        snapshot[n] = (
            _num(ws.cell(r, LEHUI_COL_T).value),
            _num(ws.cell(r, LEHUI_COL_U).value),
        )

    # 按姓名写
    written = {}
    for name, row in LEHUI_ROWS.items():
        if name not in data:
            # 这次没给 → 保留原值（不去清零）
            written[name] = {"orders": snapshot[name][0], "amount": snapshot[name][1],
                             "status": "未更新-沿用原值"}
            continue
        orders, amount = data[name]
        ws.cell(row, LEHUI_COL_T).value = _num(orders)
        ws.cell(row, LEHUI_COL_U).value = _num(amount)
        written[name] = {"orders": _num(orders), "amount": _num(amount),
                         "status": "已更新"}

    # 保护行头与公式行不动
    # T12=乐回收标题, T13=单量表头, U13=增值表头, T19/U19=SUM 公式
    # openpyxl 不会因 .value=None 而清掉公式（T19/U19 本就是字符串 "=SUM(...)"），只要别显式 set None 就好。

    wb.save(xlsx)

    if verify:
        # 重新打开校验公式未失活 + 写入生效
        wb2 = openpyxl.load_workbook(xlsx, data_only=False)
        ws2 = wb2[LEHUI_SHEET]
        for r in (19,):
            for c in (LEHUI_COL_T, LEHUI_COL_U):
                cell = ws2.cell(r, c)
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    sys.exit(f"❌ T19/U19 公式被改坏了: {cell.coordinate} = {cell.value!r}")
    return written


def parse_kv_args(argv_data):
    """从 argv 里取 --邵乐乐 8 138 风格的参数。
    argv_data: list of (name, orders, amount) tuples 来自 parse_known_args 的 unknown。
    """
    out = {}
    if len(argv_data) % 3 != 0:
        sys.exit("❌ --<姓名> 形式必须是三件套：--姓名 单量 增值")
    for i in range(0, len(argv_data), 3):
        name = argv_data[i].lstrip("-").strip()
        if name not in LEHUI_ROWS:
            sys.exit(f"❌ 姓名 [{name}] 不在固定单元格映射里: {list(LEHUI_ROWS)}")
        try:
            orders = float(argv_data[i + 1])
            amount = float(argv_data[i + 2])
        except ValueError:
            sys.exit(f"❌ {name} 单量/增值必须是数字：{argv_data[i+1]!r}, {argv_data[i+2]!r}")
        out[name] = (orders, amount)
    return out


def main():
    ap = argparse.ArgumentParser(description="乐回收更新工具（写入李家村销售表 T/U 列）")
    ap.add_argument("--xlsx", required=True, help="李家村8月任务进度.xlsx 路径")
    ap.add_argument("--data", help="JSON：{姓名: [单量, 增值], ...}")
    ap.add_argument("--no-verify", action="store_true",
                    help="跳过公式完整性校验（不推荐）")
    args, unknown = ap.parse_known_args()

    data = {}
    if args.data:
        try:
            raw = json.loads(args.data)
        except json.JSONDecodeError as e:
            sys.exit(f"❌ --data 不是合法 JSON: {e}")
        for name, val in raw.items():
            if name not in LEHUI_ROWS:
                sys.exit(f"❌ 姓名 [{name}] 不在固定映射: {list(LEHUI_ROWS)}")
            if not (isinstance(val, (list, tuple)) and len(val) == 2):
                sys.exit(f"❌ {name} 必须是 [单量, 增值] 两元素数组: {val!r}")
            data[name] = (val[0], val[1])

    kv = parse_kv_args(unknown)
    for k, v in kv.items():
        data[k] = v

    if not data:
        sys.exit("❌ 没拿到任何数据：传 --data 'JSON' 或 --姓名 单量 增值")

    print("📝 写入计划：")
    for n, (o, a) in data.items():
        print(f"   {n:6s} (行{LEHUI_ROWS[n]:2d})  单量={o:.0f}  增值={a:.0f}")
    print()
    print(f"📂 表: {args.xlsx}")
    print(f"⚠️  未在 data 里的姓名将保留原值（不会被清零）")
    print()

    res = update_lehuishou(args.xlsx, data, verify=not args.no_verify)
    print("✅ 写入完成：")
    total_orders = total_amount = 0.0
    for n, r in LEHUI_ROWS.items():
        v = res.get(n, {"orders": 0.0, "amount": 0.0, "status": "?"})
        flag = "🆕" if v["status"] == "已更新" else "  "
        print(f"   {flag} {n:6s}  单量 {v['orders']:>6.0f}  公司净利 {v['amount']:>7.0f}"
              f"  [{v['status']}]")
        if v["status"] == "已更新":
            total_orders += v["orders"]
            total_amount += v["amount"]
    print(f"\n📊 本次合计：{total_orders:.0f} 单 / 公司净利 {total_amount:.0f} 元")
    print(f"   (T19/U19 SUM 公式已自动汇总)")
    print(f"\n🕓 {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
